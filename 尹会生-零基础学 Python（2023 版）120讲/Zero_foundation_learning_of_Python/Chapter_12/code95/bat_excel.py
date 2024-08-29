from pathlib import PurePath, Path

import openpyxl

P = Path('.')
files = [x for x in P.iterdir() if PurePath(x).match("*.xlsx")]
print(files)

example_xlsx = openpyxl.Workbook()
example_sheet = example_xlsx.active

for file in files:
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            print(cell.value)
            if type(cell.value) == list or type(cell.value) == tuple or type(cell.value) == dict:
                example_sheet.append(cell.value)
            else:
                example_sheet.append([cell.value])

    # 获取全部工作簿
    print(wb.sheetnames)
    # 获取单个工作簿
    print(wb['Sheet'])

example_xlsx.save('example.xlsx')
